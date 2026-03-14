# -*- coding: utf-8 -*-
import datetime
import os
from itertools import product
import sys
import psutil
import gc
import numpy as np
import pandas as pd
import xarray as xr
from openpyxl import Workbook

from message_ix_buildings.chilled.functions.buildings_funcs_grid import (
    P_f,
    Q_c_tmax,
    Q_c_tmax_d,
    Q_h,
    calc_E_c_ac,
    calc_E_c_fan,
    calc_E_c_fan_d,
    calc_E_h,
    calc_gn_sol,
    calc_gn_sol_h,
    calc_gn_sol_tot,
    calc_H_tr,
    calc_H_v_cl,
    calc_H_v_op,
    calc_Nd,
    calc_Nf,
    calc_SCDD_m,
    calc_SHDD_m,
    calc_t_bal_c,
    calc_t_bal_h,
    calc_t_max_c,
    calc_vdd_h,
    calc_vdd_tmax_c,
    calc_vdd_tmax_c_d,
)
from message_ix_buildings.chilled.functions.variable_dicts import (
    VARDICT_COOL,
    VARDICT_COOL_DAILY,
    VARDICT_HEAT,
    VARDICT_HEAT_DAILY,
    VARS_ARCHETYPES,
    VARUNDICT_COOL,
    VARUNDICT_COOL_DAILY,
    VARUNDICT_HEAT,
    VARUNDICT_HEAT_DAILY,
)

from message_ix_buildings.chilled.preprocess.message_raster import (
    create_message_raster,  # type: ignore
)
from message_ix_buildings.chilled.util.config import Config  # type: ignore
from message_ix_buildings.chilled.util.util import (
    get_archs,
    get_logger,
    load_all_scenarios_data,
    load_parametric_analysis_data,
)

log = get_logger(__name__, log_prefix=f"{Config.gcm}_{Config.rcp}")



import re
def find_filePath(folder,var,climate_model,scenario,year):
    #
    #
    #    
    path = []
    try:
        file_path = os.path.join(folder, var, climate_model, scenario)
        files = os.listdir(file_path)
        files.sort()
        for f in files:
            if int(re.split('_',f)[6][:4]) >= int(year[0]) and int(re.split('_',f)[6][:4]) <= int(year[1]) and re.split('_',f)[3] == scenario:
                path.append(os.path.join(file_path, f))
    except:
        file_path = os.path.join(folder, var)
        files = os.listdir(file_path)
        files.sort()
        for f in files:            
            if f.endswith('.nc') and int(re.split('_',f)[6][:4]) >= int(year[0]) and int(re.split('_',f)[6][:4]) <= int(year[1]) and re.split('_',f)[3] == scenario and re.split('_',f)[2] == climate_model:
                path.append(os.path.join(file_path, f))
    return path   



def check_file_exists(folder_path, arch_list, clim, par_index_list, config):
    missing_files = []

    for arch in arch_list:
        for par_idx in par_index_list:
            suff = f"{clim}_{arch}"
            expected_filename = f"ISO_agg_data_{suff}_{par_idx}{config.cool}{config.heat}_{config.vstr}.xlsx"
            full_path = os.path.join(folder_path, expected_filename)

            if not os.path.exists(full_path):
                missing_files.append(full_path)

    if missing_files:
        print("missing：")
        for f in missing_files:
            print(f)
        return False
    else:
        print("all files exist")
        return True


def pop_weighted_sum(var, popdata, raster): 

    popdata = popdata.chunk({"lat": 100, "lon": 144, "urt": -1})
    var = var.chunk({"lat": 100, "lon": 1440, "time": -1})
    raster = raster.chunk({"lat": 100, "lon": 144})

    df = []
    for urt_idx in range(len(popdata['urt'])):

        weighted = var * popdata.isel(urt=urt_idx)
        df.append(
            weighted.groupby(raster).sum().compute()
           )
        del weighted
        gc.collect()
        
    df = xr.concat(df, dim='urt')
    df.coords['urt'] = popdata['urt'].values
    return df    


def create_daily_climate_variables_and_iso_tables(config: "Config", start_time: datetime.datetime):
    vers_archs = get_archs(config)
    par_var = load_parametric_analysis_data(config)
    s_runs = load_all_scenarios_data(config)
    #s_runs = s_runs[s_runs['clim']>2015]
        
    # input_path = config.dle_path
    out_path = os.path.join(config.project_path, "out", "version", config.vstr)
    archetype_path = os.path.join(out_path, "rasters")
    save_path = os.path.join(out_path, "VDD_ene_calcs")

    output_path_vdd = os.path.join(
        save_path,
        config.gcm,    
        config.rcp, 
    )
    

    iso_path = os.path.join(
        out_path, 
        "iso_tables", 
        config.gcm,
        config.rcp, 
        
    )

    if not os.path.exists(output_path_vdd):
        os.makedirs(output_path_vdd)

    if not os.path.exists(iso_path):
        os.makedirs(iso_path)  
        
    if (
        config.paranalysis_mode == 0
    ):  # If running in ref mode, keep only the ref parameter set
        par_var = par_var.loc[par_var.name_run == "ref", :]
        
        
    for i in range(len(s_runs)):
        s_run = s_runs.iloc[i]
        clim = s_run.clim        
        years_clim = config.yeardic[str(clim)]  
        
        check = check_file_exists(iso_path, vers_archs, clim, list(par_var.index), config)
        
        if check:        
            continue
        else:
            # << this selects the correct years.
            # But when testing you’ll want to use just say 3 years data,
            # so set years manually, e.g.
            # years_clim = yeardic6p0[str(s_run.clim)]
            # this will be the shortcut line to make the testing faster (1 years data)
            if config.testing_mode == 1:
                years_clim = (
                    years_clim[0],
                    str(int(years_clim[0]) + 0),
                )

            nyrs_clim = int(years_clim[1]) - int(years_clim[0]) + 1 
            filepath = find_filePath(config.clim_input, 'tas', config.gcm, config.rcp, years_clim)
            print(config.gcm, config.rcp, years_clim)            


            log.info("Reading clim data")
            # concat 
            datasets = [xr.open_dataset(f) for f in filepath]
            if len(datasets)>1:
                dst = xr.concat(datasets, dim="time") 

            else:
                dst = xr.open_mfdataset(
                    filepath, 
                    combine="by_coords",   
                    #engine="netcdf4",
                    #parallel=True,         
                    )

            t_oa_gbm = dst[config.davar] - 273.15
            t_oa_gbm.coords['lon'] = (t_oa_gbm.coords['lon'] + 180) % 360 - 180
            t_oa_gbm = t_oa_gbm.sortby(t_oa_gbm.lon) 
            t_oa_gbm = t_oa_gbm.sortby(t_oa_gbm.lat, ascending = False) 
            # dates = t_oa_gbm['time'].values
            # t_oa_gbm = t_oa_gbm.transpose("lat", "lon", "time")
            # ! group by month
            t_oa_gbm = t_oa_gbm.chunk({'time': -1})
            t_oa_gbm = t_oa_gbm.groupby("time.month")  
            del dst
            gc.collect()
            log.info(datetime.datetime.now() - start_time)  
            print('read finished: clim_data',datetime.datetime.now() - start_time)

            process = psutil.Process(os.getpid())
            mem_bytes = process.memory_info().rss  # (rest set size)，unit: bytes
            mem_gb = mem_bytes / (1024 ** 3)
            print(f"Current memory usage: {mem_gb:.6f} GB")  
            
      
    
        def map_calculated_variables(args):
            arch, parset = args         
            suff = str(clim) + "_" + arch    # suffix

            # merged_final_map = []
            iso_table_name = "ISO_agg_data_" + suff + '_' + str(parset.Index) + str(Config.cool) + str(Config.heat) + "_" +  config.vstr + ".xlsx"

            # If the file exists, remove firstly and then creat a new one.
            if not os.path.exists(os.path.join(iso_path, iso_table_name)):
                # os.remove(os.path.join(iso_path, iso_table_name))

                wb = Workbook()
                wb.save(os.path.join(iso_path, iso_table_name))

                # Write all urts to the same table
                with pd.ExcelWriter( os.path.join(
                                        iso_path,
                                        iso_table_name,
                                    ), 
                                    mode="a", 
                                    engine="openpyxl", 
                                    if_sheet_exists="overlay") as writer:


                    # for urt in config.urts: 
                    log.info(str(clim) + " + " + arch + " + " + parset.name_run)
                    
                    # ------------- if we have the map, we can load directly. 
                    #               For historical periods, some repetitive calculations can be avoided. but it will not save time due to the slow read speed.
                    # file_list = [suff + "_" + str(parset.Index) + "_vdd_tmax_c_daily.nc",
                    #             suff + "_" + str(parset.Index) + "_E_c_ac_daily.nc",
                    #             suff + "_" + str(parset.Index) + "_E_c_fan_daily.nc",
                    #             suff + "_" + str(parset.Index) + "_vdd_h_daily.nc",
                    #             suff + "_" + str(parset.Index) + "_E_h_daily.nc",
                    #            ]
                    
                    try:
                        if config.cool == 1:
                            vdd_tmax_c = xr.open_dataarray(os.path.join(output_path_vdd, str(2015) + "_" + arch + "_" + str(parset.Index) + "_vdd_tmax_c_daily.nc")).load()
                            print('Read: vdd_tmax_c' )
                            E_c_ac = xr.open_dataarray(os.path.join(output_path_vdd, str(2015) + "_" + arch + "_" + str(parset.Index) + "_E_c_ac_daily.nc")).load()
                            print('Read: E_c_ac' )
                            E_c_fan = xr.open_dataarray(os.path.join(output_path_vdd, str(2015) + "_" + arch + "_" + str(parset.Index) + "_E_c_fan_daily.nc")).load()
                            print('Read: E_c_fan' )

                        if config.heat == 1:
                            vdd_h = xr.open_dataarray(os.path.join(output_path_vdd, str(2015) + "_" + arch + "_" + str(parset.Index) + "_vdd_h_daily.nc")).load()
                            print('Read: vdd_h' )
                            E_h = xr.open_dataarray(os.path.join(output_path_vdd, str(2015) + "_" + arch + "_" + str(parset.Index) + "_E_h_daily.nc")).load()
                            print('Read: E_h' )
                            
                        log.info(datetime.datetime.now() - start_time)   
                            
                    except:
                        print('No available files, start to calculate')

                        # Vertical irradiation
                        # i_sol_v = xr.open_dataarray(
                        #     os.path.join(config.dle_path, "EWEMBI_vert_irrad_1980-2009_avg.nc")
                        # ).reindex(lat=t_out_ave.lat, lon=t_out_ave.lon,method='nearest').astype("float32")  # Values  in daily Wh/m2
                        i_sol_v = xr.open_dataarray(
                            os.path.join(config.dle_path, "EWEMBI_vert_irrad_1980-2009_avg_600x1440resample.nc")
                            )

                        # Horizontal irradiation
                        # i_sol_h = xr.open_dataarray(
                        #     os.path.join(config.dle_path, "EWEMBI_horiz_irrad_1980-2009_avg.nc")
                        # ).reindex(lat=t_out_ave.lat, lon=t_out_ave.lon,method='nearest').astype("float32")  # Values in daily Wh/m2
                        i_sol_h = xr.open_dataarray(
                            os.path.join(config.dle_path, "EWEMBI_horiz_irrad_1980-2009_avg_600x1440resample.nc")
                            )                

                        # if config.arch_setting == "regional":
                        #     xr.open_dataset(
                        #         os.path.join(
                        #             archetype_path,
                        #             "arch_map_" + config.arch_setting + "_" + arch + ".nc",
                        #         )
                        #     )

                        dfa = pd.DataFrame(columns=["H_v_cl", "H_v_op", "H_tr"], index=par_var.index)


                        log.info("Starting: " + suff + "_" + str(parset.name_run))
                        if config.cool == 1:
                            cop = parset.cop
                            t_sp_c = np.int8(
                                parset.t_sp_c
                            )  # Indoor setpoint temperature for cooling -> 26
                            t_sp_c_max = np.int8(
                                parset.t_sp_c_max
                            )  # Indoor max temperature when fans are on (°C) -> 28

                            f_c = parset.f_c
                            f_f = parset.f_f


                        if config.heat == 1:
                            t_sp_h = np.int8(
                                parset.t_sp_h
                            )  # Indoor max temperature when fans are on (°C) 
                            eff = parset.eff  # Efficiency heating system

                            f_h = parset.f_h


                        def read_netcdf_files(input_args):
                            varname, arch, urt = input_args
                            var = xr.open_dataset(
                                os.path.join(
                                    archetype_path, "arch_" + arch + "_" + str(varname) + ".nc"
                                )
                            )[urt]
                            return var

                        list_args = product(VARS_ARCHETYPES, [arch], ['urban']) # there is no difference between urban and rural here
                        list_netcdf = list(map(read_netcdf_files, list_args))
                        dict_netcdf = dict(zip(VARS_ARCHETYPES, list_netcdf))

                        if config.solar_gains == "VERT":
                            # Solar gains - From windows only
                            log.info("Stage 3 - calc gn_sol")
                            gn_sol = calc_gn_sol(
                                i_sol_v,
                                dict_netcdf["gl_perc"],
                                dict_netcdf["gl_g"],
                                dict_netcdf["gl_sh"],
                            )
                            # gn_sol = gn_sol.chunk(chunks={"lon": config.chunk_size})
                            # log.info("chunked")

                            gn_sol = gn_sol.compute()

                            # save files
                            if config.gridded_maps_output:
                                gn_sol.attrs = {
                                    "name": "gn_sol",
                                    "description": "Solar gains - Windows",
                                    "units": "W/m2",
                                    "short name": "Solar gains - Windows",
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        
                                gn_sol = gn_sol.to_dataset(name="gn_sol")
                                encoding = {"gn_sol": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_gn_sol.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                gn_sol.to_netcdf(filestr, encoding=encoding)
                                log.info("...Saved " + filestr)
                                gn_sol = xr.open_dataarray(filestr).load()

                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)


                        elif config.solar_gains == "TOT":
                            # Solar gains - Total
                            log.info("Stage 3 - calc gn_sol")
                            gn_sol = calc_gn_sol_tot(
                                i_sol_v,
                                dict_netcdf["gl_perc"],
                                dict_netcdf["gl_g"],
                                dict_netcdf["gl_sh"],
                                i_sol_h,
                                dict_netcdf["roof_area"],
                                dict_netcdf["roof_abs"],
                                dict_netcdf["u_roof"],
                            )

                            gn_sol = gn_sol.compute()


                            # save file
                            if config.gridded_maps_output:
                                gn_sol.attrs = {
                                    "name": "gn_sol",
                                    "description": "Solar gains",
                                    "units": "W/m2",
                                    "short name": "Solar gains",
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        
                                gn_sol = gn_sol.to_dataset(name="gn_sol")
                                encoding = {"gn_sol": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_gn_sol.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                gn_sol.to_netcdf(filestr, encoding=encoding)
                                log.info("...Saved " + filestr)
                                gn_sol = xr.open_dataarray(filestr).load()

                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)


                        elif config.solar_gains == "HOR":
                            # Solar gains - Total
                            log.info("Stage 3 - calc gn_sol")
                            gn_sol = calc_gn_sol_h(
                                i_sol_h,
                                dict_netcdf["roof_area"],
                                dict_netcdf["roof_abs"],
                                dict_netcdf["u_roof"],
                            )
                            # gn_sol = gn_sol.chunk(chunks={"lon": config.chunk_size})
                            # log.info("chunked")

                            gn_sol = gn_sol.compute()

                            # save file
                            if config.gridded_maps_output:

                                gn_sol.attrs = {
                                    "name": "gn_sol",
                                    "description": "Solar gains",
                                    "units": "W/m2",
                                    "short name": "Solar gains",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        
                                gn_sol = gn_sol.to_dataset(name="gn_sol")
                                encoding = {"gn_sol": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_gn_sol.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                gn_sol.to_netcdf(filestr, encoding=encoding)
                                log.info("...Saved " + filestr)
                                gn_sol = xr.open_dataarray(filestr).load()

                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)



                        # ==============================================================================
                        # Heat transfer functions
                        # ==============================================================================

                        H_v_cl = calc_H_v_cl(dict_netcdf["vol"], dict_netcdf["ach_cl"]).compute()
                        if config.verbose:
                            log.info("Stage 3 - calc_H_v_cl")
                            log.info(datetime.datetime.now() - start_time)
                        # H_v_cl = xr.open_dataarray(output_folder2+'H_v_cl_'+urt+'.nc').load()


                        H_v_op = calc_H_v_op(dict_netcdf["vol"], dict_netcdf["ach_op"]).compute()
                        if config.verbose:
                            log.info("Stage 3 - calc_H_v_op")
                            log.info(datetime.datetime.now() - start_time)
                        # H_v_op = xr.open_dataarray(output_folder2+'H_v_op_'+urt+'.nc').load()


                        H_tr = calc_H_tr(dict_netcdf["u_val"], dict_netcdf["area_env"]).compute()
                        if config.verbose:
                            log.info("Stage 3 - calc_H_tr")
                            log.info(datetime.datetime.now() - start_time)
                        #    H_tr = xr.open_dataarray(output_folder2+'H_tr_'+urt+'.nc').load()
                        dfa.loc[parset.Index, :] = [H_v_cl, H_v_op, H_tr]


                        if config.cool == 1:
                            # ==============================================================================
                            # Variable CDD functions
                            # ==============================================================================
                            log.info("t_bal_c")
                            t_bal_c = calc_t_bal_c(
                                t_sp_c, dict_netcdf["gn_int"], gn_sol, H_tr, H_v_cl
                            ).astype("float32")           
                            # t_bal_c = t_bal_c.chunk(chunks={"lon": config.chunk_size})
                            # log.info("chunked")
                            # t_bal_c = t_bal_c.compute()

                            # save file
                            if config.gridded_maps_output:
                                t_bal_c.attrs = {
                                    "name": "t_bal_c",
                                    "description": "Balance (base) temperature",
                                    "units": "degC",
                                    "short name": "Balance temp.",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        
                                t_bal_c = t_bal_c.to_dataset(
                                    name="t_bal_c"
                                )  # comment out because already a Dataset                
                                encoding = {"t_bal_c": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_t_bal_c.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                t_bal_c.to_netcdf(filestr, encoding=encoding)
                                log.info("...Saved " + filestr)
                                t_bal_c = xr.open_dataarray(filestr)

                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)


                            # =============================================================================
                            # t_max_c
                            # =============================================================================
                            log.info("Calc_t_max_c")
                            t_max_c = calc_t_max_c(
                                t_sp_c_max, dict_netcdf["gn_int"], gn_sol, H_tr, H_v_op
                            ).astype("float32")  # , x_diff0)
                            # t_max_c = t_max_c.compute()

                            # save file
                            if config.gridded_maps_output:
                                t_max_c.attrs = {
                                    "name": "t_max_c",
                                    "description": "This returns the max temperature",
                                    "units": "degC",
                                    "short name": "Max temp.",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        

                                t_max_c = t_max_c.to_dataset(name="t_max_c")
                                encoding = {"t_max_c": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_t_max_c.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                t_max_c.to_netcdf(filestr, encoding=encoding)
                                t_max_c = t_max_c['E_c_fan']
                                log.info("...Saved " + filestr)
                         

                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)


                            # =============================================================================
                            # vdd_tmax_c
                            # =============================================================================
                            # degreeDays over t_max_c                  
                            log.info("Calc_vdd_tmax_c")
                            vdd_tmax_c = calc_vdd_tmax_c_d(t_oa_gbm, t_max_c, t_bal_c)
                            vdd_tmax_c = vdd_tmax_c.chunk({"time": -1})                  
                            vdd_tmax_c = vdd_tmax_c.compute().astype("float32")
                            process = psutil.Process(os.getpid())
                            mem_bytes = process.memory_info().rss  # (rest set size), unit: bytes
                            mem_gb = mem_bytes / (1024 ** 3)

                            print(f"Step vdd_c, Current memory usage: {mem_gb:.6f} GB")

                            ## save file
                            if config.daily_gridded_maps_output:
                                vdd_tmax_c.attrs = {
                                    "name": "vdd_tmax_c",
                                    "description": "This returns the sum of variable cooling degree days based on Tmax",
                                    "units": "degC",
                                    "short name": "Var. cooling DD",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        
                                vdd_tmax_c = vdd_tmax_c.to_dataset(name="vdd_tmax_c")
                                encoding = {"vdd_tmax_c": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_vdd_tmax_c_daily.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                vdd_tmax_c.to_netcdf(filestr, encoding=encoding)
                                vdd_tmax_c = vdd_tmax_c['vdd_tmax_c']
                                log.info("...Saved " + filestr)
                           

                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)


                            # =============================================================================
                            # qctmax
                            # =============================================================================
                            # t_bal_c = xr.open_dataarray(
                            #     os.path.join(
                            #         output_path_vdd,
                            #         suff + "_" + str(parset.Index) + "_t_bal_c_" + urt + ".nc",
                            #     )
                            # )

                            log.info("Calc_qctmax")
                            qctmax = Q_c_tmax_d(H_tr, H_v_cl, vdd_tmax_c, f_c).astype("float32") 
                            qctmax = qctmax.compute()

                            # qctmax = qctmax.to_dataset(name="qctmax")
                            # ## save file
                            if config.gridded_maps_output:
                                qctmax.attrs = {
                                    "name": "qctmax",
                                    "description": "This returns the monthly cooling energy (MJ) based on variable degree days",
                                    "units": "MJ",
                                    "short name": "Sensible load",
                                    "AC_hours": str(f_c),
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }

                                encoding = {"qctmax": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_qctmax.nc" 
                                filestr = os.path.join(output_path_vdd, fname)
                                qctmax.to_netcdf(filestr, encoding=encoding)
                                log.info("...Saved " + filestr)
                                qctmax = xr.open_dataarray(filestr)

                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)

                            process = psutil.Process(os.getpid())
                            mem_bytes = process.memory_info().rss  # (rest set size), unit bytes
                            mem_gb = mem_bytes / (1024 ** 3)
                            print(f"Step qctmax, Current memory usage: {mem_gb:.6f} GB")    

                            # =============================================================================
                            # E_c_ac electricity
                            # =============================================================================           

                            log.info("E_c AC")
                            E_c_ac = calc_E_c_ac(qctmax, cop).astype("float32")
                            E_c_ac = E_c_ac.compute()


                            if config.daily_gridded_maps_output:    
                                E_c_ac.attrs = {
                                    "name": "E_c_ac",
                                    "description": "electricity requirement for air conditioning (MJ)",
                                    "units": "MJ/day",
                                    "short name": "AC energy sens",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }

                                E_c_ac = E_c_ac.to_dataset(name="E_c_ac")
                                encoding = {"E_c_ac": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_E_c_ac_daily.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                E_c_ac.to_netcdf(filestr, encoding=encoding)
                                E_c_ac = E_c_ac['E_c_ac']
                                log.info("...Saved " + filestr)


                            if config.monthly_gridded_maps_output:
                                E_c_ac_m = E_c_ac.groupby("month").sum(dim="time")
                                E_c_ac_m.attrs = {
                                    "name": "E_c_ac",
                                    "description": "electricity requirement for air conditioning (MJ)",
                                    "units": "MJ/m2/day",
                                    "short name": "AC energy sens",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }   
                                if not config.daily_gridded_maps_output:
                                    E_c_ac_m = E_c_ac_m.to_dataset(name="E_c_ac")
                                encoding = {"E_c_ac": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_E_c_ac_monthly.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                E_c_ac_m.to_netcdf(filestr, encoding=encoding)                                  
                                del E_c_ac_m
                                gc.collect()


                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)


                            log.info("E_f fans")               
                            E_c_fan = calc_E_c_fan_d(
                                f_f, P_f, t_oa_gbm, t_bal_c, t_max_c, config.area_fan
                            ).astype("float32").compute()  # Where Nf is same as Nd
                            E_c_fan = E_c_fan.rename("E_c_fan")


                            if config.daily_gridded_maps_output:
                                E_c_fan.attrs = {
                                    "name": "E_c_fan",
                                    "description": "daily electricity requirement for fans (MJ)",
                                    "units": "MJ/m2/day",
                                    "short name": "Fan energy",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        

                                E_c_fan = E_c_fan.to_dataset(name="E_c_fan")
                                encoding = {"E_c_fan": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_E_c_fan_daily.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                E_c_fan.to_netcdf(filestr, encoding=encoding)
                                E_c_fan = E_c_fan['E_c_fan']
                                log.info("...Saved " + filestr)


                            if config.monthly_gridded_maps_output:
                                E_c_fan_m = E_c_fan.groupby("month").sum(dim="time")
                                E_c_fan_m.attrs = {
                                    "name": "E_c_fan",
                                    "description": "daily electricity requirement for fans (MJ)",
                                    "units": "MJ/m2/day",
                                    "short name": "Fan energy",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }          
                                if not config.daily_gridded_maps_output:
                                    E_c_fan_m = E_c_fan_m.to_dataset(name="E_c_fan")
                                encoding = {"E_c_fan": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_E_c_fan_monthly.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                E_c_fan_m.to_netcdf(filestr, encoding=encoding)                            
                                del E_c_fan_m
                                gc.collect()



                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time) 

                            del qctmax
                            gc.collect()

                            process = psutil.Process(os.getpid())
                            mem_bytes = process.memory_info().rss  # rest set size, bytes
                            mem_gb = mem_bytes / (1024 ** 3)
                            print(f"Step E_c_ac+fan, Current memory usage: {mem_gb:.6f} GB")

                        # ==============================================================================
                        # HEATING CALCULATIONS
                        # ==============================================================================

                        if config.heat == 1:
                            # ==============================================================================
                            # Variable HDD functions
                            # ==============================================================================
                            log.info("calc_t_bal_h")
                            t_bal_h = calc_t_bal_h(
                                t_sp_h, dict_netcdf["gn_int"], gn_sol, H_tr, H_v_cl
                            ).astype("float32")  
                            # t_bal_h = t_bal_h.chunk({"time": -1})                    
                            # log.info("chunked")
                            t_bal_h = t_bal_h.compute()


                            if config.gridded_maps_output:
                                t_bal_h.attrs = {
                                    "name": "t_bal_h",
                                    "description": "Balance (base) temperature",
                                    "units": "degC",
                                    "short name": "Balance temp.",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        

                                t_bal_h = t_bal_h.to_dataset(name="t_bal_h")
                                encoding = {"t_bal_h": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_t_bal_h.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                t_bal_h.to_netcdf(filestr, encoding=encoding)
                                log.info("...Saved " + filestr)
                                t_bal_h = xr.open_dataarray(filestr)

                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)


                            # =============================================================================
                            # vdd_h
                            # =============================================================================
                            log.info("calc_vdd_h")
                            vdd_h = calc_vdd_h(t_oa_gbm, t_bal_h)
                            vdd_h = vdd_h.chunk({"time": -1})
                            # vdd_h = (
                            #     vdd_h.groupby("time.month").sum("time") / nyrs_clim
                            # )  # <<< divide by years                  
                            vdd_h = vdd_h.compute().astype("float32")

                            if config.daily_gridded_maps_output:
                                vdd_h.attrs = {
                                    "name": "vdd_h",
                                    "description": "This returns the variable heating degree days per day",
                                    "units": "degC",
                                    "short name": "Var. heating DD",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        
                                vdd_h = vdd_h.to_dataset(name="vdd_h")
                                encoding = {"vdd_h": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_vdd_h_daily.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                vdd_h.to_netcdf(filestr, encoding=encoding)
                                vdd_h = vdd_h['vdd_h']
                                log.info("...Saved " + filestr)
                                


                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)


                            # =============================================================================
                            # qh
                            # =============================================================================
                            # t_bal_h = xr.open_dataarray(
                            #     os.path.join(
                            #         output_path_vdd,
                            #         suff + "_" + str(parset.Index) + "_t_bal_h_" + urt + ".nc",
                            #     )
                            # )
                            log.info("Calc_qh")
                            qh = Q_h(H_tr, H_v_cl, f_h, vdd_h).astype("float32")
                            # qh = qh.compute()  

                            if config.gridded_maps_output:
                                qh.attrs = {
                                    "name": "qh",
                                    "description": "This returns the daily heating energy (MJ) based on variable degree days",
                                    "units": "MJ/day",
                                    "short name": "Sensible load",
                                    "heating_hours": str(f_h),
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        

                                qh = qh.to_dataset(name="qh")
                                encoding = {"qh": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_qh.nc" 
                                filestr = os.path.join(output_path_vdd, fname)
                                qh.to_netcdf(filestr, encoding=encoding)
                                log.info("...Saved " + filestr)

                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)
                            # qh = xr.open_dataarray(filestr)

                            # =============================================================================
                            # E_h final energy
                            # =============================================================================
                            # qlat_month = xr.open_dataarray(output_folder2+'qlat_month_'+urt+'.nc')
                            log.info("E_h")
                            E_h = calc_E_h(qh, eff).astype("float32")
                            # E_h = E_h.compute()  


                            if config.daily_gridded_maps_output:
                                E_h.attrs = {
                                    "name": "E_h",
                                    "description": "daily final energy for heating (MJ)",
                                    "units": "MJ/day",
                                    "short name": "heating energy",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }                        
                                E_h = E_h.to_dataset(name="E_h")
                                encoding = {"E_h": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_E_h_daily.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                E_h.to_netcdf(filestr, encoding=encoding)
                                E_h = E_h['E_h']

                            if config.monthly_gridded_maps_output:
                                E_h_m = E_h.groupby("month").sum(dim="time")
                                E_h_m.attrs = {
                                    "name": "E_h",
                                    "description": "monthly final energy for heating (MJ)",
                                    "units": "MJ/m2/month",
                                    "short name": "heating energy",
                                    #"urt": urt,
                                    "name_run": parset.name_run,
                                    "id_run": str(parset.Index),
                                }       
                                if not config.daily_gridded_maps_output:
                                    E_h_m = E_h_m.to_dataset(name="E_h")
                                encoding = {"E_h": config.comp}
                                fname = suff + "_" + str(parset.Index) + "_E_h_monthly.nc"
                                filestr = os.path.join(output_path_vdd, fname)
                                E_h_m.to_netcdf(filestr, encoding=encoding)                            
                                del E_h_m
                                gc.collect()


                            if config.verbose:
                                log.info(datetime.datetime.now() - start_time)
                            #    qlat_month = xr.open_dataarray(output_folder2+'qlat_month_'+urt+'.nc')
                            #dfa.to_csv(
                            #    os.path.join(output_path_vdd, suff + "_" + "constant_vars_out.csv")
                            #)
                            # log.info('Finished!: '+suff+'+str(parset))
                            log.info('Finished!')
                            log.info(datetime.datetime.now() - start_time)   

                            del qh
                            gc.collect()

                            process = psutil.Process(os.getpid())
                            mem_bytes = process.memory_info().rss  # rest set size, bytes
                            mem_gb = mem_bytes / (1024 ** 3)
                            print(f"Step E_h, Current memory usage: {mem_gb:.6f} GB")                           

                        
                        

                    def process_final_maps(args):
                        arch, parset = args

                        input_path = config.dle_path
                        out_path = os.path.join(config.project_path, "out", "version", config.vstr)
                        vdd_path = os.path.join(
                            out_path,
                            "VDD_ene_calcs",
                            config.gcm,           
                            config.rcp,     
                        )


                                     
                        vardic = {}
                        varundic = {}

                        # Update dictionaries if config.cool == 1
                        if config.cool == 1:
                            vardic.update(VARDICT_COOL_DAILY)
                            varundic.update(VARUNDICT_COOL_DAILY)
                        # Update dictionaries if config.heat == 1
                        if config.heat == 1:
                            vardic.update(VARDICT_HEAT_DAILY)
                            varundic.update(VARUNDICT_HEAT_DAILY)

                        varlist = [key for key in vardic.keys()]


                        # Load population data
                        popdata = []
                        for urt in config.urts:                            
                            if config.popfix is True:
                                suff2 = "ssp2"  + "_" + urt + "_"  + str(s_run.year)
                            else:
                                suff2 = (
                                    str(s_run.scen) + "_" + urt + "_" + str(s_run.year)
                                )  
                            
                            popdata.append( xr.open_dataarray(
                                os.path.join(
                                    input_path, "population", suff2  + ".nc"
                                )))
                            
                        popdata = xr.concat(popdata, dim=pd.Index(config.urts, name='urt'))
                     


                        # =============================================================================
                        # First apply population weighting to energy demand:
                        # =============================================================================

                        final_map = (
                            xr.Dataset()
                        )    
                     
                        log.info("process_climate_zone_tables")
                        clim_zone = xr.open_dataset(os.path.join(input_path, "climate_zones_600x1440resample.nc"))
                        clim_zone_dict = {int(k): v for k, v in clim_zone.attrs.items() if k.isdigit()}   
                        clim_zone_dict.update({99:'NAN'})
                        
                        
                        log.info("process_iso_tables")
                        # Read raster data
                        country_map = xr.open_dataarray(os.path.join(input_path, "gaul_lvl0_hybrid_05_3_600x1440.nc"))
                        country_dict = country_map.attrs.copy()
                        country_dict.update({'-1':'NAN'})
                        
                        
                        encode_factor = 100  # climatic zone: 0-26, country_id: 0-184
                        group_id = (country_map.astype(np.int32) * encode_factor + clim_zone['combined'].astype(np.int32))
                        group_id.values[group_id < 0] = -1
                        
                        

                        if config.cool == 1:            
                            log.info("final_map_E_c")
                            
                            # Cooling outputs
                            # E_c = E_c_ac + E_c_fan
                            # final_map['E_c_perpix'] = (
                            #     floorarea[urt] * E_c
                            # ).astype(np.float32)  # energy  per person per pixel

                            final_map['E_c_ac_popwei'] = pop_weighted_sum(E_c_ac, popdata, group_id)
                            gc.collect()

                            #(
                            #    pop_value  * E_c_ac #* floorarea[urt]
                            #)  # Total Energy, population weighted, per pixel

                            final_map['E_c_fan_popwei'] = pop_weighted_sum(E_c_fan, popdata, group_id)
                            gc.collect()
                            #(
                            #    pop_value  * E_c_fan#* floorarea[urt]
                            #)  # Total Energy, population weighted, per pixel

                            final_map['E_c_popwei'] = final_map['E_c_ac_popwei'] + final_map['E_c_fan_popwei']
                            gc.collect()
                            
                            #(
                            #    final_map['E_c_fan_popwei'] + final_map['E_c_ac_popwei']
                            #)  # Total Energy, population weighted, per pixel'

                            final_map['vdd_c_popwei'] = pop_weighted_sum(vdd_tmax_c, popdata, group_id)
                            gc.collect()
                            
                            #(
                            #    pop_value * vdd_tmax_c
                            #)  # Degree Days multiplied by population
                            log.info(datetime.datetime.now() - start_time) 


                        if config.heat == 1:    
                            # Heating results      
                            log.info("final_map_E_h")
                            final_map['E_h_popwei'] = pop_weighted_sum(E_h, popdata, group_id)
                            gc.collect()
                            #(
                            #    pop_value  * E_h #* floorarea[urt]
                            #)  # Total Energy, population weighted, per pixel


                            final_map['vdd_h_popwei'] = pop_weighted_sum(vdd_h, popdata, group_id)
                            gc.collect()
                            #(
                            #    pop_value * vdd_h
                            #)  # Degree Days multiplied by population
                            log.info(datetime.datetime.now() - start_time)
                            

                            
                            

                        # Read country data
                        dfd = pd.read_csv(
                            os.path.join(input_path, "GAUL_lvl0_raster0.5.csv"), index_col="ID"
                        ).assign(ISONUM=lambda x: x.index)

                        # Import MESSAGE regions and North/South classification
                        msgNS = pd.read_excel(
                            os.path.join(input_path, "country_data_" + config.vstrcntry + ".xlsx"),
                            sheet_name="ssp2_2010",
                        )

                        # Add 'GLOBAL_SOUTH' and 'REGION_GEA' to dfd
                        dfd = dfd.merge(
                            msgNS.reindex(columns=["ISO", "GLOBAL_SOUTH", "REGION_GEA"]),
                            left_on="ISO3",
                            right_on="ISO",
                        ).set_index("ISONUM")



                        # TODO: (meas) the original code does not query for clims,
                        # but without it the code will crash if not all years have been run
                        # clims_int = list(map(int, config.clims))
                        # log.info("Years of data available: " + str(clims_int))
                        # s_runs = s_runs.query("clim in @clims_int")          

                        # agg grid to country scale
                        agg_popdata = (  
                                        popdata.groupby(group_id).sum().astype(np.float32)                        
                                       )
                   


                        #agg_ras_daily = (
                        #    final_map.groupby(group_id).sum()
                        #)


                        agg_ras_yearly = (
                            final_map.sum('time').to_dataframe().reset_index()
                        )      
                       
                        agg_ras_yearly[list(final_map.data_vars.keys())] = agg_ras_yearly[list(final_map.data_vars.keys())] / nyrs_clim
                        
                        # decode
                        agg_ras_yearly['gaul_lvl0'] = agg_ras_yearly['group'] // encode_factor
                        agg_ras_yearly['clim_zone_id'] = agg_ras_yearly['group'] % encode_factor
                        agg_ras_yearly['clim_zone'] = [clim_zone_dict[i] for i in agg_ras_yearly['clim_zone_id']]


                        final_map['vdd_c_avg'] = final_map['vdd_c_popwei'] / agg_popdata
                        final_map['vdd_h_avg'] = final_map['vdd_h_popwei'] / agg_popdata
                        final_map['E_c_ac_avg'] = final_map['E_c_ac_popwei'] / agg_popdata
                        final_map['E_c_fan_avg'] = final_map['E_c_fan_popwei'] / agg_popdata
                        final_map['E_c_avg'] = final_map['E_c_popwei'] / agg_popdata
                        final_map['E_h_avg'] = final_map['E_h_popwei'] / agg_popdata


                        # decode
                        gaul_lvl0 = final_map.group.values // encode_factor
                        climate_zone_id = final_map.group.values % encode_factor
                        country_code = [country_dict[str(i)] for i in gaul_lvl0]
                        climate_code = [clim_zone_dict[i] for i in climate_zone_id]
                        final_map.coords['gaul_lvl0'] = ('group', gaul_lvl0)
                        final_map.coords['country_iso'] = ('group', country_code)
                        final_map.coords['clim_zone_id'] = ('group', climate_zone_id)
                        final_map.coords['clim_zone_name'] = ('group', climate_code)


                        final_map.attrs = {
                            "description": "Daily energy demand",
                            "units": "MJ/m2/day",
                            "short name": "cooling and heating energy",
                            "name_run": parset.name_run,
                            "id_run": str(parset.Index),
                        }

                        # to save sapce:
                        final_map = final_map.drop_vars(['E_c_ac_popwei', 'E_c_fan_popwei', 'vdd_c_popwei', 'vdd_h_popwei'])
                        
                        encoding = {var_name: config.comp for var_name in list(final_map.data_vars)}
                        fname = "ISO_agg_daily_" + suff + '_' + str(parset.Index) + str(Config.cool) + str(Config.heat) + "_" +  config.vstr + ".nc"
                        filestr = os.path.join(iso_path, fname)
                        final_map.to_netcdf(filestr, encoding=encoding)
                        log.info("Completed aggregating daily raster data!"+ str(datetime.datetime.now()))
                        
                        process = psutil.Process(os.getpid())
                        mem_bytes = process.memory_info().rss  # rest set size, bytes
                        mem_gb = mem_bytes / (1024 ** 3)
                        print(f"Step final map, Current memory usage: {mem_gb:.6f} GB")       

                        

                        df_agg = agg_ras_yearly.assign(
                                gcm=str(config.gcm),
                                scenario=str(config.rcp),
                                scen=str(s_run.scen),
                                year=str(s_run.year),
                                clim=str(s_run.clim),
                                arch=str(arch),
                                #urt=str(urt),
                                par_var=str(parset.Index),
                                name_run=str(parset.name_run),
                            #varname=str_varname,
                        ).merge(dfd, left_on="gaul_lvl0", right_on="ISONUM")
                    



                        # Load population data
                        agg_popdata = (
                             agg_popdata
                            .to_dataframe()
                            .reset_index()
                            #.melt(id_vars="gaul_lvl0", var_name="urt", value_name="popsum")
                            .assign(population_scenario=s_run.scen, year=s_run.year)
                        )
                        agg_popdata['gaul_lvl0'] = agg_popdata['group'] // encode_factor
                        agg_popdata['clim_zone_id'] = agg_popdata['group'] % encode_factor
                        agg_popdata['clim_zone'] = [clim_zone_dict[i] for i in agg_popdata['clim_zone_id']]     
                        
                        

                        # Merge df_agg with pop_agg
                        df_agg = df_agg.assign(year=lambda x: x.year.astype(int)).merge(
                            agg_popdata,
                            left_on=["gaul_lvl0",'clim_zone_id','clim_zone','year','urt'],
                            right_on=["gaul_lvl0",'clim_zone_id','clim_zone','year','urt'],
                        )


                        # in case "ZeroDivisionError: float division by zero"
                        df_agg.loc[df_agg['population'] == 0, "population"] = 1

                        # Calculate population-averaged degree days
                        if config.cool == 1:
                            df_agg = df_agg.assign(vdd_c_avg=lambda x: x.vdd_c_popwei / x.population)
                            df_agg = df_agg.assign(E_c_ac_avg=lambda x: x.E_c_ac_popwei / x.population)
                            df_agg = df_agg.assign(E_c_fan_avg=lambda x: x.E_c_fan_popwei / x.population)
                            df_agg = df_agg.assign(E_c_avg=lambda x: x.E_c_popwei / x.population)

                        if config.heat == 1:
                            df_agg = df_agg.assign(vdd_h_avg=lambda x: x.vdd_h_popwei / x.population)
                            df_agg = df_agg.assign(E_h_avg=lambda x: x.E_h_popwei / x.population)
                            
              



                        # reorder the columns
                        order = [
                                "gaul_lvl0",
                                "ADM0_CODE",
                                "ADM0_NAME",
                                "CONTINENT",
                                "FAO_CODE",
                                "ISO3",
                                "UN_CODE",
                                "UN_REGION",
                                "ISO",
                                "GLOBAL_SOUTH",
                                "REGION_GEA",
                                "gcm",
                                "scenario",
                                "scen",
                                "population_scenario",
                                "year",
                                "clim",
                                'clim_zone_id',
                                'clim_zone',
                                "arch",
                                "urt",
                                "par_var",
                                "name_run",
                                "population",
                                # 'time',
                                ] + varlist 

                        df_agg = df_agg[order]
                       

                        # Drop and rename columns
                        df_agg = df_agg.drop(
                            columns=["ADM0_CODE", "CONTINENT", "FAO_CODE", "ISO3", "UN_CODE", "UN_REGION"]
                        ).rename(columns={"gaul_lvl0": "id", "ADM0_NAME": "NAME"})


                        # Save to excel                    
                        # If it is written for the first time (startrow == 1, 
                        # indicating only the table header), 
                        # the table header is retained; otherwise, the table header is removed
                        startrow = 0 if writer.sheets["Sheet"].max_row == 1 else writer.sheets["Sheet"].max_row
                        df_agg.to_excel(writer, 
                                        sheet_name = "Sheet", 
                                        startrow = startrow, 
                                        index = False, 
                                        header = (startrow == 0)
                                        )


            
                        log.info(
                            "Done! Total time to aggregate variables and process ISO tables: "
                            + str(datetime.datetime.now())
                        )


                    inputs = (arch, parset)
                    process_final_maps(inputs) 
                    # del final_map, vdd_h, E_h, E_c_ac, E_c_fan, vdd_tmax_c
                    gc.collect()                    
                    # print('finished:', s_run.clim, arch, parset.name_run, urt, datetime.datetime.now())   


        inputs = product(vers_archs, par_var.itertuples())
        list(map(map_calculated_variables, inputs))  
        del t_oa_gbm
        gc.collect()
    

